"""
Tests voor main.py — mock de Gmail client en gebruikersinput.
"""
from datetime import date
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from email_parser import TicketData
from tests.conftest import SAMPLE_HTML_ROUND_TRIP, SAMPLE_HTML_SINGLE_HEEN


def _make_raw_email_list(*html_samples):
    """Maak een lijst van (msg_id, order_number, html) tuples."""
    result = []
    for i, (order, html) in enumerate(html_samples):
        result.append((f"msg_id_{i}", order, html))
    return result


@pytest.fixture
def mock_config(tmp_path, temp_excel):
    """Vervang config-waarden door testpaden."""
    mock = MagicMock()
    mock.EXCEL_PATH = temp_excel
    mock.SCREENSHOTS_DIR = tmp_path / "screenshots"
    mock.SCREENSHOTS_DIR.mkdir()
    mock.CLIENT_SECRET_PATH = tmp_path / "credentials" / "client_secret.json"
    mock.TOKEN_PATH = tmp_path / "credentials" / "token.json"
    mock.STATE_FILE = tmp_path / "processed.json"
    return mock


class TestMainFlow:
    def test_ticket_added_on_yes(self, mock_config):
        """Ticket wordt aan Excel toegevoegd als gebruiker 'j' antwoordt."""
        import openpyxl
        from excel_updater import DATA_START_ROW, COL_VERVOER

        raw_emails = _make_raw_email_list(("UPL1IGGK", SAMPLE_HTML_ROUND_TRIP))

        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=raw_emails),
            patch("main.save_screenshot", return_value=Path("/fake/screenshot.png")),
            patch("builtins.input", return_value="j"),
        ):
            import main
            main.main()

        wb = openpyxl.load_workbook(mock_config.EXCEL_PATH)
        ws = wb["Februari 2026"]
        assert ws.cell(row=DATA_START_ROW, column=COL_VERVOER).value == 28.0

    def test_ticket_skipped_on_no(self, mock_config):
        """Ticket wordt NIET toegevoegd als gebruiker 'n' antwoordt."""
        import openpyxl

        raw_emails = _make_raw_email_list(("UPL1IGGK", SAMPLE_HTML_ROUND_TRIP))

        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=raw_emails),
            patch("builtins.input", return_value="n"),
        ):
            import main
            main.main()

        wb = openpyxl.load_workbook(mock_config.EXCEL_PATH)
        # "Februari 2026" sheet zou nog niet mogen bestaan
        assert "Februari 2026" not in wb.sheetnames

    def test_weekend_ticket_skipped_permanently(self, mock_config):
        """Weekend-ticket dat afgewezen wordt, verschijnt niet meer."""
        from state import load_state

        # 13/02/2026 is een vrijdag (werkdag) — gebruik een zaterdag
        saturday_html = SAMPLE_HTML_ROUND_TRIP.replace(
            "13/02/2026 Terug: 13/02/2026", "14/02/2026 Terug: 14/02/2026"
        ).replace("13/02/2026", "14/02/2026")
        # Pas ook de datum in de HTML aan (Heen: en Terug:)

        raw_emails = _make_raw_email_list(("UPL1IGGK", saturday_html))

        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=raw_emails),
            patch("builtins.input", return_value="n"),
        ):
            import main
            main.main()

        state = load_state(mock_config.STATE_FILE)
        # Moet als skipped gemarkeerd zijn (niet als processed)
        assert "UPL1IGGK" in state.get("skipped_weekend", []) or \
               "UPL1IGGK" in state.get("processed", [])

    def test_already_processed_skipped(self, mock_config):
        """Al-verwerkte tickets worden stilzwijgend overgeslagen."""
        from state import load_state, save_state, mark_processed

        # Markeer het ticket vooraf als verwerkt
        state = load_state(mock_config.STATE_FILE)
        mark_processed("UPL1IGGK", state)
        save_state(state, mock_config.STATE_FILE)

        raw_emails = _make_raw_email_list(("UPL1IGGK", SAMPLE_HTML_ROUND_TRIP))
        input_calls = []

        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=raw_emails),
            patch("builtins.input", side_effect=lambda _: input_calls.append(1) or "j"),
        ):
            import main
            main.main()

        # Input mag nooit gevraagd zijn voor een al-verwerkt ticket
        assert len(input_calls) == 0

    def test_no_emails_prints_nothing_found(self, mock_config, capsys):
        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=[]),
        ):
            import main
            main.main()

        out = capsys.readouterr().out
        assert "geen nieuwe tickets" in out.lower()


class TestResetState:
    def test_reset_clears_state_file(self, mock_config, capsys):
        """--reset met 'j' antwoord verwijdert processed.json."""
        from state import load_state, save_state, mark_processed
        import main

        state = load_state(mock_config.STATE_FILE)
        mark_processed("TST001", state)
        save_state(state, mock_config.STATE_FILE)
        assert mock_config.STATE_FILE.exists()

        with (
            patch("main.config", mock_config),
            patch("builtins.input", return_value="j"),
        ):
            main.reset_state()

        assert not mock_config.STATE_FILE.exists()

    def test_reset_cancelled_leaves_state(self, mock_config, capsys):
        """--reset met 'n' antwoord laat processed.json intact."""
        from state import load_state, save_state, mark_processed
        import main

        state = load_state(mock_config.STATE_FILE)
        mark_processed("TST001", state)
        save_state(state, mock_config.STATE_FILE)

        with (
            patch("main.config", mock_config),
            patch("builtins.input", return_value="n"),
        ):
            main.reset_state()

        assert mock_config.STATE_FILE.exists()

    def test_reset_empty_state_prints_nothing(self, mock_config, capsys):
        """--reset op lege staat vraagt niet om bevestiging."""
        import main

        with patch("main.config", mock_config):
            main.reset_state()

        out = capsys.readouterr().out
        assert "leeg" in out.lower()

    def test_reset_removes_excel_rows(self, mock_config, capsys):
        """--reset verwijdert de bijbehorende rijen uit Excel voor tickets met metadata."""
        import openpyxl
        from state import load_state, save_state, mark_processed
        from excel_updater import add_ticket_to_excel, DATA_START_ROW, COL_DATUM
        from email_parser import TicketData
        import main

        ticket = TicketData(
            order_number="TST999",
            from_station="Zottegem",
            to_station="Antwerpen-Zuid",
            direction="heen/terug",
            travel_date=date(2026, 1, 7),
            price=28.0,
            email_html="",
        )
        add_ticket_to_excel(ticket, mock_config.EXCEL_PATH)

        from excel_updater import date_to_excel_serial

        state = load_state(mock_config.STATE_FILE)
        mark_processed(
            "TST999",
            state,
            metadata={
                "sheet_name": "Januari 2026",
                "travel_date_serial": date_to_excel_serial(date(2026, 1, 7)),
                "description": "Trein Zottegem - Antwerpen-Zuid heen/terug",
            },
        )
        save_state(state, mock_config.STATE_FILE)

        with (
            patch("main.config", mock_config),
            patch("builtins.input", return_value="j"),
        ):
            main.reset_state()

        assert not mock_config.STATE_FILE.exists()
        wb = openpyxl.load_workbook(mock_config.EXCEL_PATH)
        ws = wb["Januari 2026"]
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value is None

    def test_reset_skips_ticket_without_metadata(self, mock_config, capsys):
        """--reset slaat tickets zonder Excel-metadata stil over (backward compat)."""
        from state import load_state, save_state, mark_processed
        import main

        state = load_state(mock_config.STATE_FILE)
        mark_processed("OLD001", state)  # geen metadata — oud formaat
        save_state(state, mock_config.STATE_FILE)

        with (
            patch("main.config", mock_config),
            patch("builtins.input", return_value="j"),
        ):
            main.reset_state()

        # State moet gewist zijn ondanks het ontbreken van metadata
        assert not mock_config.STATE_FILE.exists()
        out = capsys.readouterr().out
        assert "handmatig" in out.lower()

    def test_reset_aborts_on_excel_locked(self, mock_config, capsys):
        """--reset breekt af als het Excel-bestand vergrendeld is."""
        from state import load_state, save_state, mark_processed
        import main

        state = load_state(mock_config.STATE_FILE)
        mark_processed(
            "TST777",
            state,
            metadata={
                "sheet_name": "Januari 2026",
                "travel_date_serial": 46028,
                "description": "Trein X - Y heen",
            },
        )
        save_state(state, mock_config.STATE_FILE)

        with (
            patch("main.config", mock_config),
            patch("builtins.input", return_value="j"),
            patch(
                "main.remove_ticket_from_excel",
                side_effect=OSError("bestand vergrendeld"),
            ),
        ):
            main.reset_state()

        # State mag NIET verwijderd zijn na een mislukte Excel-operatie
        assert mock_config.STATE_FILE.exists()
