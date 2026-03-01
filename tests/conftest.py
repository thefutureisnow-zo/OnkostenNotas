"""
Gedeelde testfixtures.
"""
import calendar
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from constants import DUTCH_MONTHS
from excel_updater import date_to_excel_serial

# ---------------------------------------------------------------------------
# Sample e-mail HTML (gebaseerd op echte NMBS-bevestigingsmail UPL1IGGK)
# ---------------------------------------------------------------------------

SAMPLE_HTML_ROUND_TRIP = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>NMBS E-Ticket</title></head>
<body>
  <p>Dag STIJN VAN DER SPIEGEL,</p>
  <table width="100%"><tr><td>
    <table width="100%" cellpadding="0" cellspacing="0" border="0"
           style="border-bottom: 1px solid gray; margin-top: 20px;">
      <tr><td style="font-weight: bold; font-size: 18px;">
        Bestelnummer: UPL1IGGK
      </td></tr>
    </table>
  </td></tr></table>
  <div><span>Van : </span><span style="font-weight: bold;">ZOTTEGEM</span></div>
  <div><span>Naar : </span><span style="font-weight: bold;">ANTWERPEN-ZUID</span></div>
  <div>2e klas, Heen en terug</div>
  <div>Heen: 13/02/2026 Terug: 13/02/2026</div>
  <table><tr>
    <td>Totaalbedrag :</td>
    <td>&euro; 28,00</td>
  </tr></table>
</body>
</html>"""

SAMPLE_HTML_SINGLE_HEEN = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>NMBS E-Ticket</title></head>
<body>
  <p>Dag STIJN VAN DER SPIEGEL,</p>
  <table width="100%"><tr><td>
    <table width="100%" cellpadding="0" cellspacing="0" border="0"
           style="border-bottom: 1px solid gray; margin-top: 20px;">
      <tr><td style="font-weight: bold; font-size: 18px;">
        Bestelnummer: ABC12345
      </td></tr>
    </table>
  </td></tr></table>
  <div><span>Van : </span><span style="font-weight: bold;">ZOTTEGEM</span></div>
  <div><span>Naar : </span><span style="font-weight: bold;">ANTWERPEN-ZUID</span></div>
  <div>2e klas, Enkel</div>
  <div>Heen: 07/01/2026</div>
  <table><tr>
    <td>Totaalbedrag :</td>
    <td>&euro; 14,00</td>
  </tr></table>
</body>
</html>"""

SAMPLE_HTML_SINGLE_TERUG = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>NMBS E-Ticket</title></head>
<body>
  <p>Dag STIJN VAN DER SPIEGEL,</p>
  <table width="100%"><tr><td>
    <table width="100%" cellpadding="0" cellspacing="0" border="0"
           style="border-bottom: 1px solid gray; margin-top: 20px;">
      <tr><td style="font-weight: bold; font-size: 18px;">
        Bestelnummer: XYZ99999
      </td></tr>
    </table>
  </td></tr></table>
  <div><span>Van : </span><span style="font-weight: bold;">ANTWERPEN-ZUID</span></div>
  <div><span>Naar : </span><span style="font-weight: bold;">ZOTTEGEM</span></div>
  <div>2e klas, Enkel</div>
  <div>Terug: 07/01/2026</div>
  <table><tr>
    <td>Totaalbedrag :</td>
    <td>&euro; 14,00</td>
  </tr></table>
</body>
</html>"""

# Reproduceert de echte bug: NMBS-label zegt "Heen:" maar de stations
# (Antwerpen-Zuid -> Zottegem) tonen dat het een terugrit is.
SAMPLE_HTML_WRONG_LABEL = """<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>NMBS E-Ticket</title></head>
<body>
  <p>Dag STIJN VAN DER SPIEGEL,</p>
  <table width="100%"><tr><td>
    <table width="100%" cellpadding="0" cellspacing="0" border="0"
           style="border-bottom: 1px solid gray; margin-top: 20px;">
      <tr><td style="font-weight: bold; font-size: 18px;">
        Bestelnummer: WR826GNF
      </td></tr>
    </table>
  </td></tr></table>
  <div><span>Van : </span><span style="font-weight: bold;">ANTWERPEN-ZUID</span></div>
  <div><span>Naar : </span><span style="font-weight: bold;">ZOTTEGEM</span></div>
  <div>2e klas, Enkel</div>
  <div>Heen: 04/02/2026</div>
  <table><tr>
    <td>Totaalbedrag :</td>
    <td>&euro; 14,00</td>
  </tr></table>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Minimale Excel-fixture die de structuur van de echte onkostennota nabootst
# ---------------------------------------------------------------------------

_date_to_excel_serial = date_to_excel_serial


def _build_minimal_excel(path: Path, existing_rows: int = 0) -> Path:
    """
    Maak een minimaal Excel-bestand dat de structuur van de onkostennota nabootst.
    `existing_rows` rijen worden vooraf ingevuld (voor overflow-tests).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Januari 2026"

    # Koptekst
    ws["A4"] = "Naam"
    ws["B4"] = "Stijn Van der Spiegel"
    ws["A5"] = "Maand"
    ws["B5"] = " Januari 2026"
    ws["J4"] = "Van"
    ws["K4"] = _date_to_excel_serial(date(2026, 1, 1))
    ws["J5"] = "Tot"
    ws["K5"] = _date_to_excel_serial(date(2026, 1, 31))

    # Kolomkoppen (rij 7)
    headers = ["Datum", "Nr", "Omschrijving van de kosten", "Curr.",
               "Brandstof", "Vervoer", "Beurs", "Maaltijden",
               "Parking", "Materiaal", "Diversen", "Tot. EUR"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col).value = h

    # Datarijen met SOM-formule in kolom L
    for row in range(8, 16):
        ws.cell(row=row, column=12).value = f"=SUM(E{row}:K{row})"

    # Vul vooraf bestaande rijen in (voor overflow-tests)
    for i in range(existing_rows):
        row = 8 + i
        ws.cell(row=row, column=1).value = _date_to_excel_serial(date(2026, 1, i + 1))
        ws.cell(row=row, column=2).value = i + 1
        ws.cell(row=row, column=3).value = f"Trein Zottegem - Antwerpen-Zuid heen"
        ws.cell(row=row, column=4).value = "EUR"
        ws.cell(row=row, column=6).value = 14.0

    # Samenvattingsrijen
    ws.cell(row=16, column=6).value = "=SUM(F8:F15)"
    ws.cell(row=17, column=11).value = "Subtotaal"
    ws.cell(row=17, column=12).value = "=SUM(L8:L15)"
    ws.cell(row=18, column=1).value = "Fietsvergoeding"
    ws.cell(row=18, column=11).value = "Voorschotten"
    ws.cell(row=19, column=11).value = "TOTAAL"
    ws.cell(row=19, column=12).value = "=(L17-L18)"
    ws.cell(row=20, column=1).value = "Goedgekeurd"

    wb.save(path)
    return path


@pytest.fixture
def sample_html_round_trip():
    return SAMPLE_HTML_ROUND_TRIP


@pytest.fixture
def sample_html_single_heen():
    return SAMPLE_HTML_SINGLE_HEEN


@pytest.fixture
def sample_html_single_terug():
    return SAMPLE_HTML_SINGLE_TERUG


@pytest.fixture
def temp_excel(tmp_path):
    """Leeg Excel-bestand klaar voor tests."""
    return _build_minimal_excel(tmp_path / "test_onkosten.xlsx", existing_rows=0)


@pytest.fixture
def temp_excel_full(tmp_path):
    """Excel-bestand met alle 8 datarijen al ingevuld (voor overflow-test)."""
    return _build_minimal_excel(tmp_path / "test_onkosten_full.xlsx", existing_rows=8)
