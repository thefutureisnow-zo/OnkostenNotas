"""
Tests voor de --month maandfilter en per-maand Excel-bestanden.
"""
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from constants import DUTCH_MONTHS, DUTCH_MONTHS_REVERSE
from email_parser import TicketData
from datetime import datetime

from excel_updater import (
    excel_path_for_date,
    add_ticket_to_excel,
    remove_ticket_from_excel,
    date_to_excel_serial,
    _is_date_cell,
    DATA_START_ROW,
    COL_DATUM,
    COL_NR,
    COL_OMSCHRIJVING,
    COL_VERVOER,
    COL_TOTAAL,
)


def _make_ticket(
    order="TST00001",
    from_s="Zottegem",
    to_s="Antwerpen-Zuid",
    direction="heen",
    travel_date=date(2026, 1, 7),
    price=14.0,
):
    return TicketData(
        order_number=order,
        from_station=from_s,
        to_station=to_s,
        direction=direction,
        travel_date=travel_date,
        price=price,
        email_html="",
    )


# ---------------------------------------------------------------------------
# DUTCH_MONTHS_REVERSE
# ---------------------------------------------------------------------------


class TestDutchMonthsReverse:
    def test_all_months_present(self):
        assert len(DUTCH_MONTHS_REVERSE) == 12

    def test_lowercase_lookup(self):
        assert DUTCH_MONTHS_REVERSE["januari"] == 1
        assert DUTCH_MONTHS_REVERSE["december"] == 12

    def test_case_insensitive_key(self):
        # Keys are stored lowercase
        assert "januari" in DUTCH_MONTHS_REVERSE
        assert "Januari" not in DUTCH_MONTHS_REVERSE


# ---------------------------------------------------------------------------
# parse_month_arg
# ---------------------------------------------------------------------------


class TestParseMonthArg:
    def test_month_only(self):
        from main import parse_month_arg

        month, year = parse_month_arg("januari")
        assert month == 1
        assert year == 2026  # current year (test assumes 2026)

    def test_month_with_year(self):
        from main import parse_month_arg

        month, year = parse_month_arg("februari 2025")
        assert month == 2
        assert year == 2025

    def test_case_insensitive(self):
        from main import parse_month_arg

        month, year = parse_month_arg("MAART")
        assert month == 3

    def test_invalid_month_raises(self):
        from main import parse_month_arg

        with pytest.raises(SystemExit):
            parse_month_arg("foobar")

    def test_all_months_parseable(self):
        from main import parse_month_arg

        for name in DUTCH_MONTHS.values():
            month, _ = parse_month_arg(name.lower())
            assert 1 <= month <= 12


# ---------------------------------------------------------------------------
# excel_path_for_date
# ---------------------------------------------------------------------------


class TestExcelPathForDate:
    def test_returns_correct_filename(self, tmp_path):
        result = excel_path_for_date(tmp_path, date(2026, 1, 15))
        assert result == tmp_path / "Onkosten_Januari_2026.xlsx"

    def test_february(self, tmp_path):
        result = excel_path_for_date(tmp_path, date(2026, 2, 1))
        assert result == tmp_path / "Onkosten_Februari_2026.xlsx"

    def test_different_year(self, tmp_path):
        result = excel_path_for_date(tmp_path, date(2025, 12, 3))
        assert result == tmp_path / "Onkosten_December_2025.xlsx"


# ---------------------------------------------------------------------------
# Per-month Excel creation + add_ticket
# ---------------------------------------------------------------------------


class TestPerMonthExcel:
    def test_creates_file_if_missing(self, tmp_path):
        """add_ticket_to_excel creates the per-month file when it doesn't exist."""
        ticket = _make_ticket(travel_date=date(2026, 1, 7))
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        result_path = add_ticket_to_excel(ticket, excel_dir)

        assert result_path.exists()
        assert result_path.name == "Onkosten_Januari_2026.xlsx"

    def test_created_file_has_correct_sheet_name(self, tmp_path):
        ticket = _make_ticket(travel_date=date(2026, 2, 4))
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        result_path = add_ticket_to_excel(ticket, excel_dir)

        wb = openpyxl.load_workbook(result_path)
        assert wb.active.title == "Februari 2026"

    def test_created_file_has_headers(self, tmp_path):
        ticket = _make_ticket(travel_date=date(2026, 1, 7))
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        result_path = add_ticket_to_excel(ticket, excel_dir)

        wb = openpyxl.load_workbook(result_path)
        ws = wb.active
        assert ws.cell(row=7, column=1).value == "Datum"
        assert ws.cell(row=7, column=3).value == "Omschrijving van de kosten"

    def test_created_file_has_date_range(self, tmp_path):
        ticket = _make_ticket(travel_date=date(2026, 2, 4))
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        result_path = add_ticket_to_excel(ticket, excel_dir)

        wb = openpyxl.load_workbook(result_path)
        ws = wb.active
        # K4 = first day of month, K5 = last day (datetime na opslaan+laden)
        assert ws["K4"].value == datetime(2026, 2, 1)
        assert ws["K5"].value == datetime(2026, 2, 28)

    def test_ticket_data_written(self, tmp_path):
        ticket = _make_ticket(travel_date=date(2026, 1, 7), price=14.0)
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 7))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value == datetime(2026, 1, 7)
        assert ws.cell(row=DATA_START_ROW, column=COL_NR).value == 1
        assert "Zottegem" in ws.cell(row=DATA_START_ROW, column=COL_OMSCHRIJVING).value
        assert ws.cell(row=DATA_START_ROW, column=COL_VERVOER).value == 14.0

    def test_multiple_tickets_same_month(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        for i in range(3):
            ticket = _make_ticket(
                order=f"TST{i:04d}",
                travel_date=date(2026, 1, i + 1),
                price=14.0,
            )
            add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_NR).value == 1
        assert ws.cell(row=DATA_START_ROW + 1, column=COL_NR).value == 2
        assert ws.cell(row=DATA_START_ROW + 2, column=COL_NR).value == 3

    def test_different_months_create_separate_files(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        add_ticket_to_excel(
            _make_ticket(order="JAN001", travel_date=date(2026, 1, 5)), excel_dir
        )
        add_ticket_to_excel(
            _make_ticket(order="FEB001", travel_date=date(2026, 2, 10)), excel_dir
        )

        jan_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        feb_path = excel_path_for_date(excel_dir, date(2026, 2, 1))
        assert jan_path.exists()
        assert feb_path.exists()
        assert jan_path != feb_path

    def test_sum_formula_present(self, tmp_path):
        ticket = _make_ticket(travel_date=date(2026, 1, 7))
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 7))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        formula = ws.cell(row=DATA_START_ROW, column=COL_TOTAAL).value
        assert formula is not None
        assert "SUM" in str(formula).upper()


# ---------------------------------------------------------------------------
# Per-month Excel removal
# ---------------------------------------------------------------------------


class TestPerMonthRemove:
    def test_remove_from_per_month_file(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        ticket = _make_ticket(travel_date=date(2026, 1, 7), price=14.0)
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 7))
        result = remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 7)),
            "Trein Zottegem - Antwerpen-Zuid heen",
        )
        assert result is True

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value is None

    def test_remove_returns_false_for_missing_row(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        ticket = _make_ticket(travel_date=date(2026, 1, 7))
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 7))
        result = remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 15)),
            "Trein X - Y heen",
        )
        assert result is False


# ---------------------------------------------------------------------------
# Month filter integration (main.py)
# ---------------------------------------------------------------------------


class TestMonthFilterIntegration:
    """Tests that --month filters tickets correctly in the main flow."""

    def _make_mock_config(self, tmp_path):
        from unittest.mock import MagicMock

        mock = MagicMock()
        mock.EXCEL_DIR = tmp_path / "data"
        mock.EXCEL_DIR.mkdir()
        mock.SCREENSHOTS_DIR = tmp_path / "screenshots"
        mock.SCREENSHOTS_DIR.mkdir()
        mock.CLIENT_SECRET_PATH = tmp_path / "credentials" / "client_secret.json"
        mock.TOKEN_PATH = tmp_path / "credentials" / "token.json"
        mock.STATE_FILE = tmp_path / "processed.json"
        mock.HOME_STATION = "Zottegem"
        mock.OFFICE_STATION = "Antwerpen-Zuid"
        # No EXCEL_PATH — per-month mode
        mock.spec = []
        return mock

    def test_month_filter_only_processes_matching(self, tmp_path):
        """With --month januari, only January tickets are processed."""
        from unittest.mock import patch
        from tests.conftest import SAMPLE_HTML_SINGLE_HEEN, SAMPLE_HTML_ROUND_TRIP

        mock_config = self._make_mock_config(tmp_path)

        # SAMPLE_HTML_SINGLE_HEEN has date 07/01/2026 (January)
        # SAMPLE_HTML_ROUND_TRIP has date 13/02/2026 (February)
        raw_emails = [
            ("msg1", "ABC12345", SAMPLE_HTML_SINGLE_HEEN),
            ("msg2", "UPL1IGGK", SAMPLE_HTML_ROUND_TRIP),
        ]

        with (
            patch("main.config", mock_config),
            patch("main.fetch_nmbs_emails", return_value=raw_emails),
            patch("main.save_screenshot", return_value=Path("/fake/screenshot.png")),
            patch("builtins.input", return_value="j"),
        ):
            import main

            main.main(month_filter=(1, 2026))

        # January file should exist with data
        jan_path = excel_path_for_date(mock_config.EXCEL_DIR, date(2026, 1, 1))
        assert jan_path.exists()

        # February file should NOT exist
        feb_path = excel_path_for_date(mock_config.EXCEL_DIR, date(2026, 2, 1))
        assert not feb_path.exists()
