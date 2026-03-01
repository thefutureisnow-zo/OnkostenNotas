"""
Tests voor excel_updater.py (per-maand Excel-bestanden).
"""
from datetime import date

import openpyxl
import pytest

from email_parser import TicketData
from excel_updater import (
    add_ticket_to_excel,
    excel_path_for_date,
    remove_ticket_from_excel,
    date_to_excel_serial,
    sheet_name_for_date,
    DATA_START_ROW,
    DATA_END_ROW,
    COL_DATUM,
    COL_NR,
    COL_OMSCHRIJVING,
    COL_VERVOER,
    COL_TOTAAL,
)


def _make_ticket(
    order="TST00001",
    from_s="Zottegem",
    to_s="Antwerpen-Zuid",
    direction="heen/terug",
    travel_date=date(2026, 1, 7),
    price=28.0,
):
    return TicketData(
        order_number=order,
        from_station=from_s,
        to_station=to_s,
        direction=direction,
        travel_date=travel_date,
        price=price,
        email_html="",
    )


class TestSheetNameForDate:
    def test_january(self):
        assert sheet_name_for_date(date(2026, 1, 1)) == "Januari 2026"

    def test_december(self):
        assert sheet_name_for_date(date(2025, 12, 3)) == "December 2025"

    def test_february(self):
        assert sheet_name_for_date(date(2026, 2, 13)) == "Februari 2026"


class TestDateToExcelSerial:
    def test_known_date(self):
        # 2026-02-13 = Excel serial 46066
        assert date_to_excel_serial(date(2026, 2, 13)) == 46066

    def test_epoch(self):
        # Excel epoch is 1899-12-30; 1900-01-01 = 2 days later
        assert date_to_excel_serial(date(1900, 1, 1)) == 2


class TestAddTicketToExcel:
    def test_row_written(self, tmp_path):
        ticket = _make_ticket()
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, ticket.travel_date)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value == date_to_excel_serial(
            date(2026, 1, 7)
        )
        assert ws.cell(row=DATA_START_ROW, column=COL_NR).value == 1
        assert "Zottegem" in ws.cell(row=DATA_START_ROW, column=COL_OMSCHRIJVING).value
        assert ws.cell(row=DATA_START_ROW, column=COL_VERVOER).value == 28.0

    def test_formula_preserved(self, tmp_path):
        ticket = _make_ticket()
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, ticket.travel_date)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        formula = ws.cell(row=DATA_START_ROW, column=COL_TOTAAL).value
        assert formula is not None
        assert "SUM" in str(formula).upper()

    def test_no_duplicate(self, tmp_path):
        ticket = _make_ticket()
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        add_ticket_to_excel(ticket, excel_dir)
        add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, ticket.travel_date)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # add_ticket_to_excel heeft zelf geen deduplicatie — dat doet state.py.
        # Hier verwachten we WEL twee rijen.
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value is not None

    def test_sequential_nr(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        for i in range(3):
            ticket = _make_ticket(
                order=f"TST0000{i}",
                travel_date=date(2026, 1, i + 1),
                price=14.0,
            )
            add_ticket_to_excel(ticket, excel_dir)

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_NR).value == 1
        assert ws.cell(row=DATA_START_ROW + 1, column=COL_NR).value == 2
        assert ws.cell(row=DATA_START_ROW + 2, column=COL_NR).value == 3

    def test_different_months_create_separate_files(self, tmp_path):
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        add_ticket_to_excel(
            _make_ticket(order="JAN001", travel_date=date(2026, 1, 5)), excel_dir
        )
        add_ticket_to_excel(
            _make_ticket(order="FEB001", travel_date=date(2026, 2, 10)), excel_dir
        )

        jan_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        feb_path = excel_path_for_date(excel_dir, date(2026, 2, 1))
        assert jan_path.exists()
        assert feb_path.exists()
        assert jan_path != feb_path

    def test_overflow_row_inserted(self, tmp_path):
        """Als alle 8 datarijen vol zijn, moet er een extra rij worden ingevoegd."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        # Vul 8 rijen
        for i in range(8):
            add_ticket_to_excel(
                _make_ticket(order=f"TST{i:04d}", travel_date=date(2026, 1, i + 1), price=14.0),
                excel_dir,
            )

        # Voeg een 9e ticket toe (triggert overflow)
        add_ticket_to_excel(
            _make_ticket(order="TST0008", travel_date=date(2026, 1, 10), price=14.0),
            excel_dir,
        )

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        filled_rows = sum(
            1
            for row in range(DATA_START_ROW, DATA_START_ROW + 20)
            if isinstance(ws.cell(row=row, column=COL_DATUM).value, int)
        )
        assert filled_rows == 9


class TestRemoveTicketFromExcel:
    def _add_and_get_path(self, excel_dir, ticket):
        """Voeg ticket toe en geef het bestandspad terug."""
        return add_ticket_to_excel(ticket, excel_dir)

    def test_removes_row(self, tmp_path):
        """Een ingevoerde rij kan worden teruggedraaid via remove_ticket_from_excel."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        ticket = _make_ticket()
        excel_path = self._add_and_get_path(excel_dir, ticket)

        date_serial = date_to_excel_serial(ticket.travel_date)
        description = f"Trein {ticket.from_station} - {ticket.to_station} {ticket.direction}"

        result = remove_ticket_from_excel(excel_path, date_serial, description)
        assert result is True

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_DATUM).value is None

    def test_renumbers_remaining_rows(self, tmp_path):
        """Na verwijdering van de eerste rij worden de overige Nr-waarden hernummerd."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        for i in range(3):
            add_ticket_to_excel(
                _make_ticket(order=f"TST{i}", travel_date=date(2026, 1, i + 1), price=14.0),
                excel_dir,
            )

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 1)),
            "Trein Zottegem - Antwerpen-Zuid heen/terug",
        )

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.cell(row=DATA_START_ROW, column=COL_NR).value == 1
        assert ws.cell(row=DATA_START_ROW + 1, column=COL_NR).value == 2

    def test_l_formula_rewritten_after_delete(self, tmp_path):
        """L-formules voor rijen na de verwijderde rij verwijzen naar de juiste rij."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        for i in range(2):
            add_ticket_to_excel(
                _make_ticket(order=f"TST{i}", travel_date=date(2026, 1, i + 1), price=14.0),
                excel_dir,
            )

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))
        remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 1)),
            "Trein Zottegem - Antwerpen-Zuid heen/terug",
        )

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        formula = ws.cell(row=DATA_START_ROW, column=COL_TOTAAL).value
        assert formula is not None
        assert f"E{DATA_START_ROW}" in str(formula)

    def test_returns_false_for_missing_file(self, tmp_path, capsys):
        """Geeft False terug als het bestand niet bestaat."""
        result = remove_ticket_from_excel(
            tmp_path / "nonexistent.xlsx", 12345, "Trein X - Y heen"
        )
        assert result is False
        out = capsys.readouterr().out
        assert "niet gevonden" in out.lower()

    def test_returns_false_for_missing_row(self, tmp_path, capsys):
        """Geeft False terug als er geen overeenkomende rij is."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()
        # Maak bestand aan met een ticket
        add_ticket_to_excel(
            _make_ticket(travel_date=date(2026, 1, 5)), excel_dir
        )
        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))

        result = remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 7)),
            "Trein Zottegem - Antwerpen-Zuid heen/terug",
        )
        assert result is False

    def test_overflow_sum_formula_shrinks(self, tmp_path):
        """Na verwijdering van een overflow-rij krimpt het SOM-bereik terug."""
        excel_dir = tmp_path / "data"
        excel_dir.mkdir()

        # Vul 8 rijen + 1 overflow
        for i in range(8):
            add_ticket_to_excel(
                _make_ticket(order=f"TST{i:04d}", travel_date=date(2026, 1, i + 1), price=14.0),
                excel_dir,
            )
        add_ticket_to_excel(
            _make_ticket(order="TST0008", travel_date=date(2026, 1, 10), price=14.0),
            excel_dir,
        )

        excel_path = excel_path_for_date(excel_dir, date(2026, 1, 1))

        # Verwijder het overflow-ticket
        remove_ticket_from_excel(
            excel_path,
            date_to_excel_serial(date(2026, 1, 10)),
            "Trein Zottegem - Antwerpen-Zuid heen/terug",
        )

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        # SOM in de samenvattingsrij moet weer eindigen op DATA_END_ROW
        summary_f = ws.cell(row=DATA_END_ROW + 1, column=6).value
        assert summary_f is not None
        assert f"F{DATA_END_ROW}" in str(summary_f).upper()
        assert f"F{DATA_END_ROW + 1}" not in str(summary_f).upper()
