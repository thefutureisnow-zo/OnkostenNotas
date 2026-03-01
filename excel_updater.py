"""
Voegt een verwerkt ticket toe aan het juiste per-maand Excel-bestand.

Elk maandbestand (bijv. Onkosten_Januari_2026.xlsx) bevat precies een werkblad.
Als het bestand nog niet bestaat, wordt het automatisch aangemaakt.
"""
import calendar
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from constants import DUTCH_MONTHS
from email_parser import TicketData

# Rijen waar de data in staan (inclusief)
DATA_START_ROW = 8
DATA_END_ROW = 15  # standaard 8 rijen; bij overflow wordt er een rij ingevoegd

# Kolomnummers (1-gebaseerd)
COL_DATUM = 1        # A
COL_NR = 2           # B
COL_OMSCHRIJVING = 3 # C
COL_CURR = 4         # D
COL_VERVOER = 6      # F
COL_TOTAAL = 12      # L


# Stijlen
DATE_FORMAT = "DD/MM/YYYY"
EUR_FORMAT = '#,##0.00'
THIN_SIDE = Side(style="thin")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE,
)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
TOTAAL_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TOTAAL_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True)

# Kolombreedtes (1-gebaseerd index -> breedte)
COLUMN_WIDTHS = {
    1: 14,   # A - Datum
    2: 6,    # B - Nr
    3: 42,   # C - Omschrijving
    4: 8,    # D - Curr.
    5: 12,   # E - Brandstof
    6: 12,   # F - Vervoer
    7: 12,   # G - Beurs
    8: 12,   # H - Maaltijden
    9: 12,   # I - Parking
    10: 12,  # J - Materiaal
    11: 12,  # K - Diversen
    12: 12,  # L - Tot. EUR
}


def _apply_styles(ws) -> None:
    """Past visuele opmaak toe op het werkblad.

    Detecteert automatisch de laatste gevulde datarij.
    """
    # Detecteer laatste datarij
    last_data_row = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, DATA_START_ROW + 50):
        if _is_date_cell(ws.cell(row=row, column=COL_DATUM).value):
            last_data_row = row
    if last_data_row < DATA_START_ROW:
        last_data_row = DATA_END_ROW  # lege sheet: stijl tot standaard eindrij

    # Kolombreedtes
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Naam/Maand labels (rij 4-5, kolom A) vetgedrukt
    for row in (4, 5):
        ws.cell(row=row, column=1).font = BOLD_FONT

    # Van/Tot datums (K4, K5) — datumformaat
    ws["K4"].number_format = DATE_FORMAT
    ws["K5"].number_format = DATE_FORMAT

    # Kolomkoppen (rij 7) — vetgedrukt, grijze achtergrond, rand
    for col in range(1, COL_TOTAAL + 1):
        cell = ws.cell(row=7, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Datarijen — randen, datumformaat kolom A, EUR-formaat kolommen E-L
    for row in range(DATA_START_ROW, last_data_row + 1):
        for col in range(1, COL_TOTAAL + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col == COL_DATUM:
                cell.number_format = DATE_FORMAT
            elif col >= 5:  # E t/m L: bedragen
                cell.number_format = EUR_FORMAT

    # Samenvattingsrijen zoeken en opmaken
    for row in range(last_data_row + 1, last_data_row + 10):
        k_val = ws.cell(row=row, column=11).value
        if k_val == "Subtotaal":
            # Subtotaal-rij: bovenrand, vetgedrukt
            for col in range(1, COL_TOTAAL + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=THIN_SIDE)
            ws.cell(row=row, column=11).font = BOLD_FONT
            ws.cell(row=row, column=12).font = BOLD_FONT
            ws.cell(row=row, column=12).number_format = EUR_FORMAT
        elif k_val == "TOTAAL":
            # TOTAAL-rij: blauwe achtergrond, witte tekst, vetgedrukt
            for col in range(1, COL_TOTAAL + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = TOTAAL_FILL
                cell.font = TOTAAL_FONT
                cell.border = THIN_BORDER
            ws.cell(row=row, column=12).number_format = EUR_FORMAT
        elif k_val == "Voorschotten":
            ws.cell(row=row, column=11).font = BOLD_FONT
            ws.cell(row=row, column=12).number_format = EUR_FORMAT


def sheet_name_for_date(d: date) -> str:
    return f"{DUTCH_MONTHS[d.month]} {d.year}"


def excel_path_for_date(excel_dir: Path, d: date) -> Path:
    """Geeft het pad naar het per-maand Excel-bestand voor de gegeven datum."""
    return excel_dir / f"Onkosten_{DUTCH_MONTHS[d.month]}_{d.year}.xlsx"


def date_to_excel_serial(d: date) -> int:
    """Zet een Python-datum om naar een Excel-serieel getal."""
    delta = d - date(1899, 12, 30)
    return delta.days


def _to_datetime(d: date) -> datetime:
    """Zet een Python date om naar een datetime (voor openpyxl celwaarden)."""
    return datetime(d.year, d.month, d.day)


def _is_date_cell(value) -> bool:
    """Controleert of een celwaarde een datumwaarde is (int of datetime)."""
    return isinstance(value, (int, datetime))


def _dates_match(cell_value, serial: int) -> bool:
    """Vergelijkt een celwaarde (datetime of int) met een Excel-serieel getal."""
    if isinstance(cell_value, int):
        return cell_value == serial
    if isinstance(cell_value, datetime):
        return date_to_excel_serial(cell_value.date()) == serial
    return False


def _find_next_data_row(ws) -> int:
    """Geeft het rijnummer van de eerste lege rij in het datablok.

    Scant voorbij DATA_END_ROW om al bestaande overflow-rijen te respecteren.
    Stopt zodra een lege rij of een SUM-formule wordt aangetroffen.
    """
    for row in range(DATA_START_ROW, DATA_START_ROW + 50):
        cell_value = ws.cell(row=row, column=COL_DATUM).value
        if cell_value is None:
            return row
        # Stop als we een SUM-formule raken (begin van het samenvattingsdeel)
        tot_value = ws.cell(row=row, column=COL_TOTAAL).value
        if isinstance(tot_value, str) and "SUM(" in tot_value.upper() and not _is_date_cell(cell_value):
            return row
    return DATA_START_ROW + 50  # veiligheidsgrens


def _insert_overflow_row(ws, insert_at: int) -> None:
    """
    Voegt een lege rij in voor `insert_at` en werkt de SOM-formules bij
    in het samenvattingsgedeelte zodat de nieuwe rij meegeteld wordt.
    """
    ws.insert_rows(insert_at)

    # Stel de L-formule in voor de nieuw ingevoegde rij
    ws.cell(row=insert_at, column=COL_TOTAAL).value = (
        f"=SUM(E{insert_at}:K{insert_at})"
    )

    new_last_data_row = insert_at  # dit is nu de laaste datarij

    # Werk SOM-formules bij in het samenvattingsdeel (rijen na de data)
    for summary_row in range(new_last_data_row + 1, new_last_data_row + 20):
        for col_idx in range(1, COL_TOTAAL + 1):
            cell = ws.cell(row=summary_row, column=col_idx)
            if cell.value and isinstance(cell.value, str) and "SUM(" in cell.value.upper():
                col_letter = get_column_letter(col_idx)
                # Vergroot het bereik: SUM(X8:X15) -> SUM(X8:X<new_last>)
                cell.value = re.sub(
                    rf"SUM\({col_letter}(\d+):{col_letter}(\d+)\)",
                    lambda m, c=col_letter, nl=new_last_data_row: (
                        f"SUM({c}{m.group(1)}:{c}{nl})"
                    ),
                    cell.value,
                    flags=re.IGNORECASE,
                )


def _create_month_excel(excel_path: Path, d: date) -> None:
    """
    Maak een nieuw per-maand Excel-bestand met de standaardstructuur.
    Het bestand bevat precies een werkblad met de juiste maandnaam.
    """
    sheet_name = sheet_name_for_date(d)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Koptekst
    ws["A4"] = "Naam"
    ws["B4"] = "Stijn Van der Spiegel"
    ws["A5"] = "Maand"
    ws["B5"] = f" {sheet_name}"
    ws["J4"] = "Van"
    first_day = date(d.year, d.month, 1)
    last_day = date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])
    ws["K4"] = _to_datetime(first_day)
    ws["K5"] = _to_datetime(last_day)
    ws["J5"] = "Tot"

    # Kolomkoppen (rij 7)
    headers = [
        "Datum", "Nr", "Omschrijving van de kosten", "Curr.",
        "Brandstof", "Vervoer", "Beurs", "Maaltijden",
        "Parking", "Materiaal", "Diversen", "Tot. EUR",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col).value = h

    # Datarijen met SOM-formule in kolom L
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        ws.cell(row=row, column=COL_TOTAAL).value = f"=SUM(E{row}:K{row})"

    # Samenvattingsrijen
    ws.cell(row=16, column=6).value = f"=SUM(F{DATA_START_ROW}:F{DATA_END_ROW})"
    ws.cell(row=17, column=11).value = "Subtotaal"
    ws.cell(row=17, column=12).value = f"=SUM(L{DATA_START_ROW}:L{DATA_END_ROW})"
    ws.cell(row=18, column=1).value = "Fietsvergoeding"
    ws.cell(row=18, column=11).value = "Voorschotten"
    ws.cell(row=19, column=11).value = "TOTAAL"
    ws.cell(row=19, column=12).value = "=(L17-L18)"
    ws.cell(row=20, column=1).value = "Goedgekeurd"

    _apply_styles(ws)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def remove_ticket_from_excel(
    excel_path: Path, travel_date_serial: int, description: str
) -> bool:
    """
    Verwijdert een ticketrij uit een per-maand Excel-bestand.
    Geeft True terug als de rij gevonden en verwijderd is, anders False.
    Gooit een OSError als het bestand vergrendeld is (bijv. open in Excel).
    """
    if not excel_path.exists():
        print(f"  Waarschuwing: bestand '{excel_path.name}' niet gevonden.")
        return False

    try:
        wb = openpyxl.load_workbook(excel_path)
    except PermissionError:
        raise OSError(
            f"Het Excel-bestand is vergrendeld. Sluit het eerst in Excel: {excel_path}"
        )

    ws = wb.active

    # Bepaal de laatste gevulde datarij (voor overflow-detectie)
    last_data_row = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, DATA_START_ROW + 50):
        if _is_date_cell(ws.cell(row=row, column=COL_DATUM).value):
            last_data_row = row

    # Zoek de overeenkomende rij op datum EN omschrijving
    matches = []
    for row in range(DATA_START_ROW, last_data_row + 1):
        if (
            _dates_match(ws.cell(row=row, column=COL_DATUM).value, travel_date_serial)
            and ws.cell(row=row, column=COL_OMSCHRIJVING).value == description
        ):
            matches.append(row)

    if not matches:
        print(
            f"  Waarschuwing: rij voor '{description}' niet gevonden in '{excel_path.name}'."
            " Al handmatig verwijderd?"
        )
        return False

    if len(matches) > 1:
        print(
            f"  Waarschuwing: meerdere overeenkomende rijen in '{excel_path.name}'."
            " Eerste rij wordt verwijderd."
        )
    target_row = matches[0]

    ws.delete_rows(target_row, 1)
    new_last_data_row = last_data_row - 1

    # Herschrijf per-rij L-formule voor alle datarijen op en na de verwijderde positie
    for row in range(target_row, new_last_data_row + 1):
        if _is_date_cell(ws.cell(row=row, column=COL_DATUM).value):
            ws.cell(row=row, column=COL_TOTAAL).value = f"=SUM(E{row}:K{row})"

    # Hernummer kolom B (Nr) voor alle overblijvende datarijen
    nr = 1
    for row in range(DATA_START_ROW, new_last_data_row + 1):
        if _is_date_cell(ws.cell(row=row, column=COL_DATUM).value):
            ws.cell(row=row, column=COL_NR).value = nr
            nr += 1

    # Bij overflow: verklein SOM-bereiken in de samenvattingsrijen
    if last_data_row > DATA_END_ROW:
        for summary_row in range(new_last_data_row + 1, new_last_data_row + 20):
            for col_idx in range(1, COL_TOTAAL + 1):
                cell = ws.cell(row=summary_row, column=col_idx)
                if cell.value and isinstance(cell.value, str) and "SUM(" in cell.value.upper():
                    col_letter = get_column_letter(col_idx)
                    cell.value = re.sub(
                        rf"SUM\({col_letter}(\d+):{col_letter}(\d+)\)",
                        lambda m, c=col_letter, nl=new_last_data_row: (
                            f"SUM({c}{m.group(1)}:{c}{nl})"
                        ),
                        cell.value,
                        flags=re.IGNORECASE,
                    )

    _apply_styles(ws)
    try:
        wb.save(excel_path)
    except PermissionError:
        raise OSError(
            f"Het Excel-bestand is vergrendeld. Sluit het eerst in Excel: {excel_path}"
        )
    return True


def add_ticket_to_excel(ticket: TicketData, excel_dir: Path) -> Path:
    """
    Voegt het ticket als nieuwe rij toe aan het juiste per-maand Excel-bestand.
    Maakt het bestand automatisch aan als het nog niet bestaat.
    Geeft het pad naar het bijgewerkte bestand terug.
    Gooit een OSError als het bestand vergrendeld is (bijv. open in Excel).
    """
    excel_path = excel_path_for_date(excel_dir, ticket.travel_date)

    if not excel_path.exists():
        _create_month_excel(excel_path, ticket.travel_date)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except PermissionError:
        raise OSError(
            f"Het Excel-bestand is vergrendeld. Sluit het eerst in Excel: {excel_path}"
        )

    ws = wb.active

    next_row = _find_next_data_row(ws)

    if next_row > DATA_END_ROW:
        _insert_overflow_row(ws, next_row)

    # Schrijf de ticketgegevens
    nr = next_row - DATA_START_ROW + 1
    description = (
        f"Trein {ticket.from_station} - {ticket.to_station} {ticket.direction}"
    )

    datum_cell = ws.cell(row=next_row, column=COL_DATUM)
    datum_cell.value = _to_datetime(ticket.travel_date)
    datum_cell.number_format = DATE_FORMAT
    ws.cell(row=next_row, column=COL_NR).value = nr
    ws.cell(row=next_row, column=COL_OMSCHRIJVING).value = description
    ws.cell(row=next_row, column=COL_CURR).value = "EUR"
    ws.cell(row=next_row, column=COL_VERVOER).value = ticket.price

    # Zorg dat de L-formule aanwezig is
    if ws.cell(row=next_row, column=COL_TOTAAL).value is None:
        ws.cell(row=next_row, column=COL_TOTAAL).value = (
            f"=SUM(E{next_row}:K{next_row})"
        )

    _apply_styles(ws)
    try:
        wb.save(excel_path)
    except PermissionError:
        raise OSError(
            f"Het Excel-bestand is vergrendeld. Sluit het eerst in Excel: {excel_path}"
        )
    return excel_path
